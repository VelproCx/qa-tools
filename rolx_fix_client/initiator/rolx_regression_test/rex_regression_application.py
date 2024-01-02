#!/usr/bin/python
# -*- coding: utf8 -*-
"""FIX Application"""
import argparse
import configparser
import difflib
import random
import sys
import quickfix as fix
import time
import logging
from datetime import datetime, timedelta
from model.logger import setup_logger
import json
import math
from importlib.machinery import SourceFileLoader
import os
from openpyxl import load_workbook

# 获取当前所在目录绝对路径
current_path = os.path.abspath(os.path.dirname(__file__))
# 将当前目录的路径，获取上级目录的绝对路径
Parent_path = os.path.abspath(os.path.join(current_path, "../../method"))
# 获取上级目录中一个文件的路径
generation_path = os.path.join(Parent_path, "file_generation.py")
# 获取data_comparison
data_comparison_path = os.path.join(Parent_path, "data_comparison.py")

__SOH__ = chr(1)


class Application(fix.Application):  # 定义一个类并继承‘fix.Application’类，主要用于处理收到的消息和事件

    def __init__(self, account, logger):
        # 初始化sessionID = None
        super().__init__()
        self.sessionID = None
        self.account = account
        self.logger = logger

        self.exec_id = 0
        self.orders_dict = []
        self.ptf_cancel_list = []
        self.rex_prod_bps_buy = 0.0022
        self.rol_prop_bps_sell = 0.0022
        self.result = []
        self.receive_results = []
        self.order_new = 0
        self.order_expired = 0
        self.order_accepted = 0
        self.order_rejected = 0
        self.order_filled = 0
        self.order_partially_filled = 0

    def onCreate(self, sessionID):
        # 服务器启动时候调用此方法创建
        self.sessionID = sessionID
        print(f"onCreate : Session ({sessionID.toString()})")
        return

    def onLogon(self, sessionID):
        # "客户端登陆成功时候调用此方法"
        self.sessionID = sessionID
        print(f"successful Logon to session '{sessionID.toString()}'.")
        return

    def onLogout(self, sessionID):
        self.logs_check()
        json_data = json.dumps(self.receive_results)

        module_name = "compare_field_values"
        module_path = data_comparison_path
        # 导入具有完整文件路径的模块
        module1 = SourceFileLoader(module_name, module_path).load_module()
        # 将JSON数据写入文件中
        with open('logs/recv_data.json', 'w') as file:
            file.write(json_data)
        self.result = module1.compare_field_values('../../testcases/REX_Functional_Test_Matrix.json',
                                                   'logs/recv_data.json',
                                                   'ordStatus')
        print(f"Session ({sessionID.toString()}) logout !")

        ordStatus_list = []
        errorCode_list = []
        # 循环receive_results并将value添加到列表里
        for i in self.receive_results:
            ordStatus_list.append(str(i['ordStatus']))
            if 'errorCode' in i:
                errorCode_list.append(str(i['errorCode']))

            else:
                errorCode_list.append(" ")

        # report文件里写入字段
        self.writeResExcel('report/rex_report.xlsx', ordStatus_list, 2, 'J')
        self.writeResExcel('report/rex_report.xlsx', errorCode_list, 2, 'K')
        self.writeResExcel('report/rex_report.xlsx', self.result, 2, 'L')
        return

    def toAdmin(self, message, sessionID):
        msg = message.toString().replace(__SOH__, "|")  # 特殊符号转化
        self.logger.info(f"(Core) S >> {msg}")
        return

    def toApp(self, message, sessionID):
        self.logger.info("-------------------------------------------------------------------------------------------")
        msgType = message.getHeader().getField(35)
        msg = message.toString().replace(__SOH__, "|")
        # 7.1 New Order Single
        if msgType == "D":
            orderQty = message.getField(38)
            ordType = message.getField(40)
            clOrdID = message.getField(11)
            side = message.getField(54)
            symbol = message.getField(55)
            transactTime = message.getField(60)
            if (clOrdID, orderQty, ordType,
                side, symbol, transactTime,
                ) != "":
                self.logger.info(f"(sendMsg) New Ack >> {msg}")
                self.order_new += 1
            else:
                self.logger.info(f"(sendMsg) New Ack >> {msg}" + 'New Order Single FixMsg Error!')
        # 7.4 Order Cancel Request
        elif msgType == "F":
            clOrdID = message.getField(11)
            side = message.getField(54)
            symbol = message.getField(55)
            transactTime = message.getField(60)
            if (clOrdID, side, symbol, transactTime) != "":
                self.logger.info(f"(sendMsg) Cancel Ack >> {msg}")
            else:
                self.logger.info(f"(sendMsg) Cancel Ack >> {msg}" + 'Order Cancel Request FixMsg Error!')
        return

    def fromAdmin(self, message, sessionID):
        msg = message.toString().replace(__SOH__, "|")
        self.logger.info(f"(Core) R << {msg}")
        return

    def fromApp(self, message, sessionID):
        self.logger.info("-------------------------------------------------------------------------------------------")
        # "接收业务消息时调用此方法"
        msgType = message.getHeader().getField(35)

        if msgType == 'h':
            tradingSessionID = message.getField(336)
            tradSesMode = message.getField(339)
            tradSesStatus = message.getField(340)
            msg = message.toString().replace(__SOH__, "|")
            if (tradingSessionID, tradSesMode, tradSesStatus) != '':
                self.logger.info(f"(recvMsg) Trading Session << {msg}")
            else:
                self.logger.info("(recvMsg) Trading Session Error")
        # Business Message Reject
        elif msgType == 'j':
            refSeqNum = message.getField(45)
            text = message.getField(58)
            refMsgType = message.getField(372)
            businessRejectRefID = message.getField(379)
            msg = message.toString().replace(__SOH__, "|")
            if (refSeqNum, text, refMsgType, businessRejectRefID) != '':
                self.logger.info(f"(recvMsg) Business Message << {msg}")
            else:
                self.logger.info("(recvMsg) Business Message Error")
        else:
            clOrdID = message.getField(11)
            orderID = message.getField(37)
            ordStatus = message.getField(39)
            transactTime = message.getField(60)
            fsxTransactTime = message.getField(8169)

            if ordStatus == '4':
                symbol = message.getField(55)
                print(symbol)
                if symbol == '1311':
                    new_clOrdID = int(clOrdID) + 1
                    clOrdID = str(new_clOrdID)

            # 设置匹配的阈值
            threshold = 1

            matches = difflib.get_close_matches(clOrdID, [item['clOrdId'] for item in self.receive_results], n=1,
                                                cutoff=threshold)
            if matches:
                matched_clOrdId = matches[0]
                # 拿到clOrdId去数组里循环比对
                for item in self.receive_results:
                    # 判断当前收到的消息体clOrdId是否在数组里
                    if item['clOrdId'] == matched_clOrdId:
                        # 更新该组数据的ordStatus
                        item['ordStatus'].append(str(ordStatus))
            else:
                if ordStatus != '8':
                    # 添加新的数据到数组中
                    self.receive_results.append({'clOrdId': clOrdID, 'ordStatus': [ordStatus]})

                else:
                    text = message.getField(58)
                    self.receive_results.append({'clOrdId': clOrdID, 'ordStatus': [ordStatus], 'errorCode': text})

            if msgType != '9':
                # 消息体共用tag
                avgPx = message.getField(6)
                CumQty = message.getField(14)
                exec_id = message.getField(17)
                execTransType = message.getField(20)
                orderQty = message.getField(38)
                ordType = message.getField(40)
                rule80A = message.getField(47)
                side = message.getField(54)
                symbol = message.getField(55)
                timeInForce = message.getField(59)
                clientID = message.getField(109)
                execType = message.getField(150)
                leavesQty = message.getField(151)
                cashMargin = message.getField(544)
                crossingPriceType = message.getField(8164)
                marginTransactionType = message.getField(8214)
                if symbol == '5076':
                    self.ptf_cancel_list.append(message.getField(11))
                elif symbol == '1311' or symbol == '6954':
                    self.orders_dict = message.getField(11)
                msg = message.toString().replace(__SOH__, "|")
                # 7.2 Execution Report – Order Accepted
                if ordStatus == "0":
                    execBroker = message.getField(76)
                    lastShares = message.getField(32)
                    lastPx = message.getField(31)
                    clOrdID = message.getField(11)

                    if (
                            avgPx, clOrdID, CumQty, exec_id, execTransType, lastPx, lastShares, orderID, orderQty,
                            ordType,
                            rule80A,
                            side, symbol, timeInForce, transactTime, execBroker, clientID, execType, leavesQty,
                            cashMargin,
                            crossingPriceType, fsxTransactTime, marginTransactionType) != "":
                        self.logger.info(f"(recvMsg) Order Accepted << {msg}")
                        self.order_accepted += 1
                    else:
                        self.logger.info(f"(recvMsg) Order Accepted << {msg}" + 'Order Accepted FixMsg Error!')
                    if execType != ordStatus:  # 如果不相等，记录错误日志
                        self.logger.info(
                            f"(recvMsg) Order execType error,orderStatus = {ordStatus},execType = {execType}")
                # 7.3 Execution Report – Order Rejected
                elif ordStatus == "8":
                    text = message.getField(58)  # 附加信息
                    ordRejReason = message.getField(103)  # 拒单原因
                    lastShares = message.getField(32)  # 最新成交
                    lastPx = message.getField(31)  # 最新成价
                    clOrdID = message.getField(11)
                    # 判断tag是否存在
                    if (
                            avgPx, clOrdID, CumQty, exec_id, execTransType, lastPx, lastShares, orderID, orderQty,
                            ordType,
                            rule80A,
                            side, symbol, timeInForce, transactTime, clientID, execType, leavesQty, cashMargin,
                            crossingPriceType,
                            fsxTransactTime, marginTransactionType, text, ordRejReason) != "":
                        self.logger.info(f"(recvMsg) Order Rej << {msg}")
                        self.order_rejected += 1
                    else:
                        self.logger.info(f"(recvMsg) Order Rejected << {msg}" + 'Order Rejected FixMsg Error!')
                    if execType != ordStatus:
                        self.logger.info(
                            f"(recvMsg) Order execType error,orderStatus = {ordStatus},execType = {execType}")
                # 7.6 Execution Report – Order Canceled
                elif ordStatus == "4":
                    origClOrdID = message.getField(41)
                    execBroker = message.getField(76)
                    clOrdID = message.getField(11)
                    # 判断tag是否存在
                    if (avgPx, clOrdID, CumQty, exec_id, execTransType, orderID, orderQty, ordType, rule80A,
                        side, symbol, timeInForce, transactTime, clientID, execType, leavesQty, cashMargin,
                        crossingPriceType,
                        fsxTransactTime, marginTransactionType, origClOrdID, execBroker) != "":
                        self.logger.info(f"(recvMsg) Order Canceled << {msg}")
                    else:
                        self.logger.info(f"(recvMsg) Order Canceled << {msg}" + 'Order Canceled FixMsg Error!')
                    if execType != ordStatus:
                        self.logger.info(
                            f"(recvMsg) Order execType error,orderStatus = {ordStatus},execType = {execType}")
                # 7.7 Execution Report – Trade
                elif ordStatus == "1" or ordStatus == "2":  # 部分成交
                    lastPx = float(message.getField(31))
                    lastShares = message.getField(32)
                    execBroker = message.getField(76)  # 执行经纪商
                    primaryLastPx = float(message.getField(8031))
                    routingDecisionTime = message.getField(8051)  # 通用时间
                    propExecPrice = message.getField(8165)
                    PropExecId = message.getField(8166)
                    clOrdID = message.getField(11)
                    # 公式计算期望值 FillPrice
                    adjustLastPxBuy = math.ceil(primaryLastPx * (1 + self.rex_prod_bps_buy))
                    adjustLastPxSell = math.floor(primaryLastPx * (1 - self.rex_prod_bps_sell))
                    # 判断tag是否存在
                    if (
                            avgPx, clOrdID, CumQty, exec_id, execTransType, lastPx, lastShares, orderID, orderQty,
                            ordType,
                            rule80A,
                            side, symbol, timeInForce, transactTime, execBroker, clientID, execType, leavesQty,
                            cashMargin,
                            crossingPriceType, fsxTransactTime, marginTransactionType, primaryLastPx,
                            routingDecisionTime,
                            propExecPrice, PropExecId) != "":
                        self.logger.info(
                            "(recvMsg) Order Filled << {}".format(msg) + 'Side: ' + str(
                                side) + ',' + "Fill Price: " + str(
                                lastPx) + ',' + "AdjustLastPx Of Buy: " + str(
                                adjustLastPxBuy) + ',' + "AdjustLastPx Of Sell: " + str(
                                adjustLastPxSell) + ',' + "Order Type:" + str(ordType))
                        self.logger.info("result : Order Filled ," + "ordStatus =" + ordStatus)
                        if ordStatus == "1":
                            self.order_partially_filled += 1
                        else:
                            self.order_filled += 1
                    else:
                        self.logger.info(f"(recvMsg) Order Filled << {msg}" + "Order Trade FixMsg Error!")
                    if execType != ordStatus:
                        self.logger.info(
                            f"(recvMsg) Order execType error,orderStatus = {ordStatus},execType = {execType}")
                        # Fill Price Check
                    if ordType == '1':  # 部分成交
                        if side == "1":
                            adjustLastPx = math.ceil(primaryLastPx * (1 + self.rex_prod_bps_buy))
                            # 期望值与获取的fillPrice进行比对
                            if adjustLastPx != lastPx:
                                self.logger.info(
                                    'Market Price is not matching,' + 'clOrdID：' + clOrdID + ','
                                    + 'symbol：' + symbol + ',' + 'adjustLastPx：' + str(
                                        adjustLastPx) + ',' + 'lastPx:' + str(lastPx))
                        elif side == "2":
                            adjustLastPx = math.floor(primaryLastPx * (1 - self.rex_prod_bps_sell))
                            # 期望值与获取的fillPrice进行比对
                            if adjustLastPx != lastPx:
                                self.logger.info(
                                    'Market Price is not matching,' + 'clOrdID：' + clOrdID + ',' + 'symbol：' + symbol
                                    + ',' + 'adjustLastPx：' + str(
                                        adjustLastPx) + ',' + 'lastPx:' + str(lastPx))
                #  7.8 Execution Report – End of Day Expired
                elif ordStatus == "C":  # 过期订单
                    text = message.getField(58)
                    execBroker = message.getField(76)
                    origClOrdID = message.getField(41)
                    clOrdID = message.getField(11)
                    # 判断tag是否存在
                    if (avgPx, clOrdID, CumQty, exec_id, execTransType, orderID, orderQty, ordType, rule80A,
                        side, symbol, timeInForce, transactTime, execBroker, clientID, execType, leavesQty, cashMargin,
                        crossingPriceType, fsxTransactTime, marginTransactionType, execBroker, origClOrdID, text) != "":
                        self.logger.info(f"(recvMsg) Order Expired << {msg}" + "ExpireRes = " + str(text))
                        self.logger.info("result : Order Expired ," + "ordStatus =" + ordStatus)
                        self.order_expired += 1
                    else:
                        self.logger.info(f"(recvMsg) Order Expired << {msg}" + "Order Expired FixMsg Error!")
                    if execType != ordStatus:
                        self.logger.info(
                            f"(recvMsg) Order execType error,orderStatus = {ordStatus},execType = {execType}")
            else:
                origClOrdID = message.getField(41)
                text = message.getField(58)
                cxlRejReason = message.getField(102)
                cxlRejResponseTo = message.getField(434)
                clOrdID = message.getField(11)
                msg = message.toString().replace(__SOH__, "|")
                # 判断tag是否存在
                if (clOrdID, orderID, transactTime, fsxTransactTime, origClOrdID, text,
                    cxlRejReason, cxlRejResponseTo) != "":
                    self.logger.info(f"(recvMsg) Order Canceled << {msg}" + "ordStatus = " + str(ordStatus))
                else:
                    self.logger.info(f"(recvMsg) Order Canceled << {msg}" + 'Order Canceled FixMsg Error!')

            self.onMessage(message, sessionID)
        return

    def onMessage(self, message, sessionID):
        """Processing application message here"""
        pass
        # 数据比对的方法

    # 在掉出登陆时调用logscheck方法，判断是否有这些错误打印，
    def logs_check(self):
        response = ['ps:若列表存在failed数据，请查看report.log文件']
        self.writeResExcel(
            '/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/report/rex_report.xlsx',
            response, 2, 'M')
        with open('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/logs/rex_report.log', 'r') as f:
            content = f.read()
        if 'Market Price is not matching' in content:
            self.logger.info('Market Price is NG')
            response = ['Market Price is NG']
            self.write_result_excel('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/'
                                    'report/rex_report.xlsx', response, 5, 'M')
        else:
            self.logger.info('Market Price is OK')
            response = ['Market Price is OK']
            self.write_result_excel('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/'
                                    'report/rex_report.xlsx', response, 3, 'M')
        if 'FixMsg Error' in content:
            self.logger.info('FixMsg is NG')
            response = ['FixMsg is NG']
            self.write_result_excel('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/'
                                    'report/rex_report.xlsx', response, 6, 'M')
        else:
            self.logger.info('FixMsg is OK')
            response = ['FixMsg is OK']
            self.write_result_excel('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/'
                                    'report/rex_report.xlsx', response, 4, 'M')
        if 'Order execType error' in content:
            self.logger.info("execType is NG")
            response = ['execType is NG']
            self.write_result_excel('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/'
                                    'report/rex_report.xlsx', response, 7, "M")
        else:
            self.logger.info("execType is OK")
            response = ['execType is OK']
            self.write_result_excel('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/'
                                    'report/rex_report.xlsx', response, 8, "M")

    def write_result_excel(self, filename, data, row, column):
        # 打开现有的Excel文件或者创建新的Workbook
        workbook = load_workbook(filename)  # load_workbook指打开现有文件或新的对象
        #     指定要写入数据的工作表
        sheet = workbook.active
        # 指定要写入的列号
        start_row = row  # 行号
        start_column = column  # 列号
        # 写入数据到指定列
        for row, value in enumerate(data, start=start_row):
            sheet[start_column + str(row)] = value

        # 保持并关闭工作簿
        workbook.save(filename)

    def gen_client_order_id(self):
        # "随机数生成ClOrdID"
        self.exec_id += 1
        # 获取当前时间并且进行格式转换
        t = int(time.time())
        str1 = ''.join([str(i) for i in random.sample(range(0, 9), 4)])
        return str(t) + str1 + str(self.exec_id).zfill(6)

    def insert_order_request(self, row):
        # 构造一个FIX协议消息（NewOrderSingle消息）并设置消息中的各个字段值
        msg = fix.Message()  # 创建一个空的FIX消息对象msg
        header = msg.getHeader()  # 获取消息头
        header.setField(fix.MsgType(fix.MsgType_NewOrderSingle))  # 消息类型字段（MsgType）设为NewOrderSingle（D）
        header.setField(fix.MsgType("D"))
        msg.setField(fix.Account(row["Account"]))
        msg.setField(fix.ClOrdID(self.gen_client_order_id()))
        msg.setField(fix.OrderQty(row["OrderQty"]))
        msg.setField(fix.OrdType(row["OrdType"]))
        msg.setField(fix.Side(row["Side"]))
        msg.setField(fix.Symbol(row["Symbol"]))
        # 判断订单类型,限价单读取case中的price并且设置，市价单则不设置价格
        if row["OrdType"] == "2":
            msg.setField(fix.Price(row["Price"]))  # 如果值是2，表示是限价单，要设置价格

        # 5个非必填的字段
        if row["TimeInForce"] != "":
            msg.setField(fix.TimeInForce(row["TimeInForce"]))

        if row["Rule80A"] != "":
            msg.setField(fix.Rule80A(row["Rule80A"]))

        if row["CashMargin"] != "":
            msg.setField(fix.CashMargin(row["CashMargin"]))

        # 自定义Tag
        if row["CrossingPriceType"] != "":
            msg.setField(8164, row["CrossingPriceType"])

        if row["MarginTransactionType"] != "":
            msg.setField(8214, row["MarginTransactionType"])

        # 获取TransactTime
        trstime = fix.TransactTime()
        trstime.setString(datetime.utcnow().strftime("%Y%m%d-%H:%M:%S.%f"))
        msg.setField(trstime)

        fix.Session.sendToTarget(msg, self.sessionID)  # 发送请求
        return msg  # 返回消息体

    def order_cancel_request(self, row):
        if row["Symbol"] == '5076' and row["OrdType"] == "1":
            clOrdId = self.ptf_cancel_list[0]
        elif row["Symbol"] == '5076' and row["OrdType"] == "2":
            clOrdId = self.ptf_cancel_list[1]
        else:
            clOrdId = self.orders_dict

        time.sleep(1)
        msg = fix.Message()
        header = msg.getHeader()
        header.setField(fix.MsgType(fix.MsgType_OrderCancelRequest))
        header.setField(fix.BeginString("FIX.4.2"))
        header.setField(fix.MsgType("F"))
        msg.setField(fix.OrigClOrdID(clOrdId))
        msg.setField(fix.ClOrdID(self.gen_client_order_id()))
        msg.setField(fix.Symbol(row["Symbol"]))
        msg.setField(fix.Side(row["Side"]))

        trstime = fix.TransactTime()
        trstime.setString(datetime.utcnow().strftime("%Y%m%d-%H:%M:%S.%f"))
        msg.setField(trstime)
        fix.Session.sendToTarget(msg, self.sessionID)
        return msg

    def runTestCase(self, row):
        # 判断case是下单还是取消订单
        action = row["ActionType"]
        if action == 'NewAck':
            self.insert_order_request(row)
        elif action == 'CancelAck':
            self.order_cancel_request(row)

    def load_test_case(self):
        module_name = "generation"
        module_path = generation_path
        # 导入具有完整文件路径的模块
        module1 = SourceFileLoader(module_name, module_path).load_module()

        generation = module1.generation
        """Run"""
        with open('/app/data/qa-tools/rolx_fix_client/testcases/REX_Functional_Test_Matrix.json', 'r') as f_json:
            # 生成报告模版
            generation('/app/data/qa-tools/rolx_fix_client/testcases/REX_Functional_Test_Matrix.json',
                       '/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/report/rex_report.xlsx')
            case_data_list = json.load(f_json)
            time.sleep(2)
            # 循环所有用例，并把每条用例放入runTestCase方法中，
            for row in case_data_list["testCase"]:
                if row['Id'] == "1":
                    self.insert_order_request(row)
                    time.sleep(60)
                if row["Symbol"] == '5076' and row["OrdType"] == "1" and row["Comment"] == "CancelAck":
                    time.sleep(120)
                    self.runTestCase(row)
                else:
                    time.sleep(1)
                    self.runTestCase(row)

    def read_config(self, sender, target, host, port):
        # 读取并修改配置文件
        config = configparser.ConfigParser(allow_no_value=True)
        config.optionxform = str  # 保持键的大小写
        config.read('/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/rex_regression_client.cfg')
        config.set('SESSION', 'SenderCompID', sender)
        config.set('SESSION', 'TargetCompID', target)
        config.set('SESSION', 'SocketConnectHost', host)
        config.set('SESSION', 'SocketConnectPort', port)

        with open(
                '/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/rex_regression_client.cfg', 'w'
        ) as configfile:
            config.write(configfile)


def main():
    try:
        # 使用argparse的add_argument方法进行传参
        parser = argparse.ArgumentParser()  # 创建对象
        parser.add_argument('-account', default='RSIT_ACCOUNT_1', help='choose account to use for test')
        parser.add_argument('-sender', default='RSIT_1', help='choose Sender to use for test')
        parser.add_argument('-target', default='FSX_SIT_CGW_1', help='choose Target to use for test')
        parser.add_argument('-host', default='10.4.65.1', help='choose Host to use for test')
        parser.add_argument('-port', default='5001', help='choose Port to use for test')

        # report
        setup_logger('logfix', '/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/logs/rex_report.log')
        logger = logging.getLogger('logfix')

        args = parser.parse_args()  # 解析参数
        account = args.account
        sender = args.sender
        target = args.target
        host = args.host
        port = args.port

        cfg = Application(account, logger)
        cfg.sender = sender
        cfg.target = target
        cfg.host = host
        cfg.port = port
        cfg.read_config(sender, target, host, port)
        settings = fix.SessionSettings(
            "/app/data/qa-tools/rolx_fix_client/initiator/rolx_regression_test/rex_regression_client.cfg")
        application = Application(account, logger)
        storeFactory = fix.FileStoreFactory(settings)
        logFactory = fix.FileLogFactory(settings)
        initiator = fix.SocketInitiator(application, storeFactory, settings, logFactory)

        initiator.start()
        application.load_test_case()
        # 执行完所有测试用例后等待时间
        sleep_duration = timedelta(minutes=5)
        end_time = datetime.now() + sleep_duration
        while datetime.now() < end_time:
            time.sleep(1)
        initiator.stop()

    except (fix.ConfigError, fix.RuntimeError) as e:
        print(e)
        sys.exit()


if __name__ == '__main__':
    main()
