import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side


def generation(file_path, filename):
    fields_to_remove = ['Account', 'OrderQty', 'Side', 'Price', 'TimeInForce', 'CrossingPriceType',
                        'Rule80A', 'CashMargin', 'MarginTransactionType', 'MinQty', 'OrderClassification',
                        'SelfTradePreventionId']
    with open(file_path, 'r') as f_json:
        json_data = json.load(f_json)
    for row in json_data["testCase"]:
        # 循环接收需要删除的字段
        for field in fields_to_remove:
            try:
                # 删除字段
                del row[field]
            except KeyError:
                pass

    df = pd.json_normalize(json_data["testCase"])

    df.to_excel(filename, index=False)

    workbook = load_workbook(filename)  # 替换为你的 Excel 文件路径
    # workbook = Workbook()  # 创建一个新的 Workbook

    # 选择要设置颜色的工作表
    sheet = workbook.active  # 使用默认的活动工作表
    sheet['J1'] = 'RecvOrdStatus'
    sheet['K1'] = 'RecvErrorCode'
    sheet['L1'] = 'Execution Result'
    sheet['M1'] = 'Remark'
    # 设置最合适的列宽
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

        adjusted_width = (max_length + 2) * 1.2  # 调整宽度以适应文本内容
        sheet.column_dimensions[column_letter].width = adjusted_width

    border_none = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
    for row in sheet.rows:
        for cell in row:
            cell.border = border_none

    # 指定要设置颜色的行和列
    row_number = 1
    column_start = 'A'
    column_end = 'K'
    res_column_start = 'L'
    res_column_end = 'M'

    # 设置单元格背景颜色
    fill = PatternFill(fill_type='solid', fgColor='FFFF00')
    for column in range(ord(column_start), ord(column_end) + 1):
        cell = sheet[chr(column) + str(row_number)]
        cell.fill = fill

    res_fill = PatternFill(fill_type='solid', fgColor='00FF00')
    for column in range(ord(res_column_start), ord(res_column_end) + 1):
        cell = sheet[chr(column) + str(row_number)]
        cell.fill = res_fill

    # 保存修改并关闭工作簿
    workbook.save(filename)

